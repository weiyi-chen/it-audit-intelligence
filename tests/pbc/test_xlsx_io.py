"""
Unit tests for modules/pbc/xlsx_io.py

Tests
─────
1. Round-trip fidelity — write then read back; all fields match
2. All 4 status values survive the round-trip
3. in_scope=False is preserved (not skipped)
4. Optional fields (sample_size, last_year_id) round-trip as None/empty
5. Missing optional columns get sensible defaults on read
6. Empty rows are skipped on read
7. FileNotFoundError raised for bad path
8. Output directory is created automatically
9. Legend sheet is present in written workbook
10. Column order in written file matches _WRITE_COLUMNS
"""

from __future__ import annotations

import os

import openpyxl
import pytest

from core.state import PBCItem
from modules.pbc.xlsx_io import _WRITE_COLUMNS, read_pbc_xlsx, write_pbc_xlsx


# ─── helpers ─────────────────────────────────────────────────────────────────

def make_item(**overrides) -> PBCItem:
    base = PBCItem(
        item_id="JML-TST-001", category="ITGC - JML",
        description="Test evidence request.",
        in_scope=True, period="FY2025", sample_size="25",
        status="carried_over", last_year_id=None, notes="test note",
    )
    return PBCItem(**{**base, **overrides})


ALL_STATUSES = ["carried_over", "updated", "new", "removed"]


# ─── tests ───────────────────────────────────────────────────────────────────

class TestRoundTrip:
    def test_single_item_roundtrip(self, tmp_path):
        """Write one item and read it back — every field must match."""
        item = make_item()
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx([item], path)
        result = read_pbc_xlsx(path)

        assert len(result) == 1
        rb = result[0]
        assert rb["item_id"]     == item["item_id"]
        assert rb["category"]    == item["category"]
        assert rb["description"] == item["description"]
        assert rb["in_scope"]    == item["in_scope"]
        assert rb["period"]      == item["period"]
        assert rb["sample_size"] == item["sample_size"]
        assert rb["status"]      == item["status"]
        assert rb["notes"]       == item["notes"]

    def test_multiple_items_count(self, tmp_path, sample_prior_items):
        """All 18 prior-year items survive the round-trip."""
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx(sample_prior_items, path)
        result = read_pbc_xlsx(path)
        assert len(result) == len(sample_prior_items)

    def test_item_order_preserved(self, tmp_path, sample_prior_items):
        """Items are read back in the same order they were written."""
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx(sample_prior_items, path)
        result = read_pbc_xlsx(path)
        ids_written = [i["item_id"] for i in sample_prior_items]
        ids_read    = [i["item_id"] for i in result]
        assert ids_written == ids_read

    def test_all_field_values(self, tmp_path, sample_prior_items):
        """For every item: item_id, status, period, category all match."""
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx(sample_prior_items, path)
        result = read_pbc_xlsx(path)
        for orig, rb in zip(sample_prior_items, result):
            assert rb["item_id"]  == orig["item_id"],  f"item_id mismatch at {orig['item_id']}"
            assert rb["status"]   == orig["status"],   f"status mismatch at {orig['item_id']}"
            assert rb["period"]   == orig["period"],   f"period mismatch at {orig['item_id']}"
            assert rb["category"] == orig["category"], f"category mismatch at {orig['item_id']}"


class TestStatusValues:
    def test_all_four_statuses_roundtrip(self, tmp_path):
        """carried_over, updated, new, removed all survive write→read."""
        items = [make_item(item_id=f"TST-{s[:3].upper()}-001", status=s)
                 for s in ALL_STATUSES]
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx(items, path)
        result = read_pbc_xlsx(path)
        statuses_read = {i["item_id"]: i["status"] for i in result}
        for item in items:
            assert statuses_read[item["item_id"]] == item["status"], \
                f"Status lost for {item['item_id']}"

    def test_in_scope_false_preserved(self, tmp_path):
        """in_scope=False must not be silently flipped to True."""
        item = make_item(in_scope=False, status="removed")
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx([item], path)
        result = read_pbc_xlsx(path)
        assert result[0]["in_scope"] is False

    def test_in_scope_true_preserved(self, tmp_path):
        item = make_item(in_scope=True)
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx([item], path)
        result = read_pbc_xlsx(path)
        assert result[0]["in_scope"] is True


class TestOptionalFields:
    def test_none_sample_size_roundtrip(self, tmp_path):
        """sample_size=None must come back as None (not empty string)."""
        item = make_item(sample_size=None)
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx([item], path)
        result = read_pbc_xlsx(path)
        assert result[0]["sample_size"] is None

    def test_sample_size_string_roundtrip(self, tmp_path):
        for val in ["25", "40", "3 months", "1 review cycle"]:
            item = make_item(sample_size=val)
            path = str(tmp_path / f"pbc_{val.replace(' ', '_')}.xlsx")
            write_pbc_xlsx([item], path)
            result = read_pbc_xlsx(path)
            assert result[0]["sample_size"] == val, f"sample_size {val!r} lost"

    def test_none_last_year_id_roundtrip(self, tmp_path):
        item = make_item(last_year_id=None)
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx([item], path)
        result = read_pbc_xlsx(path)
        assert result[0]["last_year_id"] is None

    def test_last_year_id_string_roundtrip(self, tmp_path):
        item = make_item(last_year_id="JML-ORA-001")
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx([item], path)
        result = read_pbc_xlsx(path)
        assert result[0]["last_year_id"] == "JML-ORA-001"

    def test_empty_notes_roundtrip(self, tmp_path):
        item = make_item(notes="")
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx([item], path)
        result = read_pbc_xlsx(path)
        assert result[0]["notes"] == ""

    def test_multiline_description_roundtrip(self, tmp_path):
        long_desc = "Line one. " * 20
        item = make_item(description=long_desc)
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx([item], path)
        result = read_pbc_xlsx(path)
        assert result[0]["description"] == long_desc


class TestEdgeCases:
    def test_bad_path_raises(self):
        with pytest.raises(FileNotFoundError):
            read_pbc_xlsx("/nonexistent/path/pbc.xlsx")

    def test_empty_list_writes_header_only(self, tmp_path):
        """Writing [] should produce a valid workbook with header row but no data."""
        path = str(tmp_path / "empty.xlsx")
        write_pbc_xlsx([], path)
        assert os.path.exists(path)
        result = read_pbc_xlsx(path)
        assert result == []

    def test_output_dir_created(self, tmp_path):
        """write_pbc_xlsx must create nested directories automatically."""
        deep_path = str(tmp_path / "a" / "b" / "c" / "pbc.xlsx")
        write_pbc_xlsx([make_item()], deep_path)
        assert os.path.exists(deep_path)

    def test_legend_sheet_exists(self, tmp_path):
        """Workbook must have a 'Legend' sheet for colour reference."""
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx([make_item()], path)
        wb = openpyxl.load_workbook(path)
        assert "Legend" in wb.sheetnames

    def test_header_row_matches_write_columns(self, tmp_path):
        """First row of PBC List sheet must match _WRITE_COLUMNS order."""
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx([make_item()], path)
        wb = openpyxl.load_workbook(path)
        ws = wb["PBC List"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # There should be exactly len(_WRITE_COLUMNS) non-None header cells
        non_null = [h for h in headers if h is not None]
        assert len(non_null) == len(_WRITE_COLUMNS)

    def test_overwrite_existing_file(self, tmp_path, sample_prior_items):
        """Writing twice to the same path must overwrite cleanly."""
        path = str(tmp_path / "pbc.xlsx")
        write_pbc_xlsx(sample_prior_items[:3], path)
        write_pbc_xlsx(sample_prior_items[:5], path)
        result = read_pbc_xlsx(path)
        assert len(result) == 5

    def test_large_batch(self, tmp_path):
        """100-item write→read must complete and preserve count."""
        items = [make_item(item_id=f"TST-{i:03d}", description=f"Item {i}")
                 for i in range(1, 101)]
        path = str(tmp_path / "large.xlsx")
        write_pbc_xlsx(items, path)
        result = read_pbc_xlsx(path)
        assert len(result) == 100

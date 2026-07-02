"""
Module B — IT Understanding Knowledge Map nodes.

Node pipeline
─────────────
  extract_entities  →  infer_relations  →  render_map

extract_entities
    Parses evidence files (IT memo, system inventory) using Claude.
    Returns List[ITEntity] — systems, persons, processes, vendors.

infer_relations
    Given the entity list + source text, Claude infers relationships
    with confidence scores.  Returns List[EntityRelationship].

render_map
    Converts entities + relationships into:
      • vis-network compatible JSON (nodes[], edges[])
      • Standalone single-file HTML with embedded network
    Sets state["map_output_html_path"] and state["network_json"].
"""

from __future__ import annotations

import json
import os
import re
from typing import Any, Dict, List

from core.llm import call_claude
from core.state import ITEntity, EntityRelationship, State

# ─────────────────────────────────────────────────────────────────────────────
# Prompts
# ─────────────────────────────────────────────────────────────────────────────

_EXTRACT_SYSTEM = """\
You are an expert IT auditor. Extract all identifiable IT entities from the
provided evidence document.

Return ONLY a valid JSON array — no markdown fences, no explanation.
Each element must conform to:
{
  "entity_id":   "<short_snake_case_id>",
  "entity_type": "<system|process|person|vendor|location|control>",
  "name":        "<display name>",
  "attributes":  { ... type-specific key-value pairs ... }
}

Attribute guidelines by type:
  system:   hosting (on-prem/cloud/managed), criticality (high/medium/low),
            owner, version, in_scope (true/false), new_this_year (true/false),
            outsourced_to (vendor name or null)
  process:  area (JML/UAR/ChangeMgmt/PrivAccess/Batch/Dev/Backup/Auth),
            system_ids (array of system entity_ids)
  person:   role, systems (array), is_key_contact (true/false)
  vendor:   services (array), soc_report_available (true/false), soc_type
  location: type (on-prem/cloud), provider
  control:  control_type, description
"""

_EXTRACT_USER = """\
Client: {client_name}
Audit Period: {audit_period}

Evidence document:
───────────────────────────────
{text}
───────────────────────────────

Extract all IT entities. Be thorough — capture every system, person with IT
responsibilities, vendor, and documented IT process.
"""

_RELATIONS_SYSTEM = """\
You are an expert IT auditor analysing relationships between IT entities.
Given a list of entities and source documentation, identify relationships.

Return ONLY a valid JSON array. Each element:
{
  "source_id":     "<entity_id>",
  "target_id":     "<entity_id>",
  "relation":      "<owns|runs_on|processes|depends_on|reviewed_by|managed_by|connects_to|hosts>",
  "confidence":    <0.0–1.0>,
  "evidence_quote":"<short quote from source text supporting this relationship>"
}

Focus on relationships relevant to IT audit: system ownership, data flows,
access management, outsourcing, and control responsibilities.
"""

_RELATIONS_USER = """\
Client: {client_name}
Audit Period: {audit_period}

Entities extracted:
{entities_json}

Source documentation:
───────────────────────────────
{text}
───────────────────────────────

Identify relationships between the entities. Minimum confidence: 0.5.
"""


# ─────────────────────────────────────────────────────────────────────────────
# Node 1 — extract_entities_node
# ─────────────────────────────────────────────────────────────────────────────

def extract_entities_node(state: State) -> dict:
    """
    Parse evidence text and scope memo to extract List[ITEntity].

    Input state fields used:
        current_year_scope_text  — scope memo / IT understanding memo
        evidence_paths           — list of file paths to parse (optional)
        client_name, audit_period

    Output state fields:
        extracted_entities       — List[ITEntity]
    """
    print("[extract_entities_node] starting")

    texts: List[str] = []

    # Prefer evidence files; fall back to scope text
    for path in state.get("evidence_paths", []):
        try:
            t = _read_text_file(path)
            if t.strip():
                texts.append(t)
        except Exception as e:
            print(f"[extract_entities_node] ⚠ could not read {path}: {e}")

    if not texts:
        scope = state.get("current_year_scope_text", "").strip()
        if scope:
            texts.append(scope)

    if not texts:
        print("[extract_entities_node] no text available — returning empty entities")
        return {"extracted_entities": []}

    combined_text = "\n\n---\n\n".join(texts)[:6000]  # token guard

    prompt = _EXTRACT_USER.format(
        client_name  = state.get("client_name", ""),
        audit_period = state.get("audit_period", ""),
        text         = combined_text,
    )

    try:
        raw = call_claude(prompt, system=_EXTRACT_SYSTEM)
        entities_raw = _parse_json_array(raw, "extract_entities_node")
        entities: List[ITEntity] = []
        for e in entities_raw:
            if not isinstance(e, dict):
                continue
            entities.append(
                ITEntity(
                    entity_id   = str(e.get("entity_id",   f"entity_{len(entities)}")),
                    entity_type = str(e.get("entity_type", "system")),
                    name        = str(e.get("name",        "Unknown")),
                    attributes  = dict(e.get("attributes", {})),
                )
            )
        print(f"[extract_entities_node] extracted {len(entities)} entities")
        return {"extracted_entities": entities}

    except Exception as exc:
        print(f"[extract_entities_node] ❌ error: {exc}")
        return {"extracted_entities": [], "error": str(exc)}


# ─────────────────────────────────────────────────────────────────────────────
# Node 2 — infer_relations_node
# ─────────────────────────────────────────────────────────────────────────────

def infer_relations_node(state: State) -> dict:
    """
    Given extracted_entities + source text, infer EntityRelationships.

    Output state fields:
        entity_relationships  — List[EntityRelationship]
    """
    print("[infer_relations_node] starting")

    entities = state.get("extracted_entities", [])
    if not entities:
        print("[infer_relations_node] no entities — skipping")
        return {"entity_relationships": []}

    texts: List[str] = []
    for path in state.get("evidence_paths", []):
        try:
            t = _read_text_file(path)
            if t.strip():
                texts.append(t)
        except Exception:
            pass
    if not texts:
        scope = state.get("current_year_scope_text", "").strip()
        if scope:
            texts.append(scope)

    combined_text = "\n\n---\n\n".join(texts)[:4000]
    entities_json = json.dumps(
        [{"entity_id": e["entity_id"], "entity_type": e["entity_type"], "name": e["name"]}
         for e in entities],
        indent=2,
    )

    prompt = _RELATIONS_USER.format(
        client_name  = state.get("client_name", ""),
        audit_period = state.get("audit_period", ""),
        entities_json = entities_json,
        text          = combined_text,
    )

    try:
        raw = call_claude(prompt, system=_RELATIONS_SYSTEM)
        rels_raw = _parse_json_array(raw, "infer_relations_node")

        valid_ids = {e["entity_id"] for e in entities}
        relations: List[EntityRelationship] = []
        for r in rels_raw:
            if not isinstance(r, dict):
                continue
            src = str(r.get("source_id", ""))
            tgt = str(r.get("target_id", ""))
            if src not in valid_ids or tgt not in valid_ids:
                continue
            conf = float(r.get("confidence", 0.5))
            if conf < 0.5:
                continue
            relations.append(
                EntityRelationship(
                    source_id    = src,
                    target_id    = tgt,
                    relation     = str(r.get("relation", "connects_to")),
                    confidence   = round(conf, 2),
                    evidence_quote = str(r.get("evidence_quote", "")),
                )
            )

        print(f"[infer_relations_node] inferred {len(relations)} relationships")
        return {"entity_relationships": relations}

    except Exception as exc:
        print(f"[infer_relations_node] ❌ error: {exc}")
        return {"entity_relationships": [], "error": str(exc)}


# ─────────────────────────────────────────────────────────────────────────────
# Node 3 — render_map_node
# ─────────────────────────────────────────────────────────────────────────────

# Color palette per entity type (matches frontend conventions)
_TYPE_COLORS: Dict[str, str] = {
    "system":   "#1e3a8a",
    "process":  "#6d28d9",
    "person":   "#0369a1",
    "vendor":   "#c2410c",
    "location": "#0f766e",
    "control":  "#14532d",
}

# Shape per entity type (vis-network shape names)
_TYPE_SHAPES: Dict[str, str] = {
    "system":   "ellipse",
    "process":  "box",
    "person":   "diamond",
    "vendor":   "star",
    "location": "database",
    "control":  "triangle",
}

# Edge color per relation
_RELATION_COLORS: Dict[str, str] = {
    "owns":        "#1d4ed8",
    "runs_on":     "#0f766e",
    "processes":   "#6d28d9",
    "depends_on":  "#b45309",
    "reviewed_by": "#14532d",
    "managed_by":  "#c2410c",
    "connects_to": "#64748b",
    "hosts":       "#0369a1",
}


def render_map_node(state: State) -> dict:
    """
    Convert entities + relationships into vis-network JSON + standalone HTML.

    Output state fields:
        map_output_html_path  — path to the generated HTML file
        network_json          — dict with {"nodes": [...], "edges": [...]}
                                ready to pass directly to the frontend
    """
    print("[render_map_node] starting")

    entities  = state.get("extracted_entities", [])
    relations = state.get("entity_relationships", [])
    out_path  = state.get("map_output_html_path", "")

    # ── Build vis-network data ────────────────────────────────────────────────
    nodes = _build_vis_nodes(entities)
    edges = _build_vis_edges(relations)
    network_json = {"nodes": nodes, "edges": edges}

    # ── Generate standalone HTML ──────────────────────────────────────────────
    html = _render_html(
        network_json = network_json,
        entities     = entities,
        client_name  = state.get("client_name", "Client"),
        audit_period = state.get("audit_period", ""),
    )

    if out_path:
        try:
            with open(out_path, "w", encoding="utf-8") as f:
                f.write(html)
            print(f"[render_map_node] HTML written to {out_path}")
        except Exception as exc:
            print(f"[render_map_node] ⚠ could not write HTML: {exc}")

    print(f"[render_map_node] {len(nodes)} nodes, {len(edges)} edges")
    return {
        "map_output_html_path": out_path,
        "network_json":         network_json,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _build_vis_nodes(entities: List[ITEntity]) -> List[Dict[str, Any]]:
    nodes = []
    for e in entities:
        et   = e.get("entity_type", "system")
        attr = e.get("attributes", {})
        size = 30 if et == "system" else 20

        # Build tooltip
        tips = [f"<b>{e['name']}</b>", f"Type: {et}"]
        for k, v in attr.items():
            if v is not None and v != "":
                tips.append(f"{k.replace('_',' ').title()}: {v}")

        nodes.append({
            "id":    e["entity_id"],
            "label": e["name"],
            "color": {"background": _TYPE_COLORS.get(et, "#64748b"),
                      "border":     "#fff",
                      "highlight":  {"background": _TYPE_COLORS.get(et, "#64748b"), "border": "#fbbf24"}},
            "font":  {"color": "#fff", "size": 12, "bold": et == "system"},
            "shape": _TYPE_SHAPES.get(et, "ellipse"),
            "size":  size,
            "title": "<br>".join(tips),
            "group": et,
            # Extra metadata surfaced in panel
            "_entity_type": et,
            "_attrs": attr,
        })
    return nodes


def _build_vis_edges(relations: List[EntityRelationship]) -> List[Dict[str, Any]]:
    edges = []
    for i, r in enumerate(relations):
        color = _RELATION_COLORS.get(r.get("relation", ""), "#94a3b8")
        edges.append({
            "id":     i,
            "from":   r["source_id"],
            "to":     r["target_id"],
            "label":  r.get("relation", ""),
            "color":  {"color": color, "highlight": "#fbbf24"},
            "width":  max(1, int(r.get("confidence", 0.7) * 3)),
            "arrows": "to",
            "title":  r.get("evidence_quote", ""),
            "font":   {"size": 10, "color": "#64748b"},
            "smooth": {"type": "curvedCW", "roundness": 0.15},
        })
    return edges


def _render_html(
    network_json: Dict[str, Any],
    entities: List[ITEntity],
    client_name: str,
    audit_period: str,
) -> str:
    """Generate a standalone HTML file with the vis-network map embedded."""
    nodes_json = json.dumps(network_json["nodes"], indent=2)
    edges_json = json.dumps(network_json["edges"], indent=2)

    # Entity type legend
    legend_items = "".join(
        f'<span style="display:inline-flex;align-items:center;gap:5px;margin-right:14px;font-size:12px">'
        f'<span style="display:inline-block;width:12px;height:12px;border-radius:50%;background:{c}"></span>'
        f'{t.title()}</span>'
        for t, c in _TYPE_COLORS.items()
        if any(e.get("entity_type") == t for e in entities)
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<title>IT Understanding Map — {client_name} {audit_period}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/vis/4.21.0/vis.min.js"></script>
<link  href="https://cdnjs.cloudflare.com/ajax/libs/vis/4.21.0/vis.min.css" rel="stylesheet"/>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:-apple-system,sans-serif;background:#f4f6fb}}
  header{{background:#1a2035;color:#fff;padding:14px 24px;display:flex;align-items:center;gap:12px}}
  header h1{{font-size:16px;font-weight:700}}
  header span{{font-size:12px;color:#94a3b8;margin-left:auto}}
  #network{{width:100%;height:calc(100vh - 90px);background:#fff}}
  .legend{{background:#fff;border-top:1px solid #e2e8f0;padding:10px 20px;display:flex;flex-wrap:wrap;align-items:center;gap:4px}}
</style>
</head>
<body>
<header>
  <h1>🗺️ IT Understanding Map</h1>
  <span>{client_name} · {audit_period}</span>
</header>
<div id="network"></div>
<div class="legend">{legend_items}</div>
<script>
const nodes = new vis.DataSet({nodes_json});
const edges = new vis.DataSet({edges_json});
const container = document.getElementById('network');
const network = new vis.Network(container, {{nodes, edges}}, {{
  nodes: {{font:{{color:'#fff'}}, borderWidth:2}},
  edges: {{font:{{align:'middle'}}}},
  physics: {{stabilization:{{iterations:150}}}},
  interaction: {{hover:true, tooltipDelay:100}},
}});
</script>
</body>
</html>"""


def _read_text_file(path: str) -> str:
    """Read text from txt/docx/pdf files."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".txt":
        with open(path, encoding="utf-8", errors="ignore") as f:
            return f.read()
    if ext == ".docx":
        try:
            from docx import Document
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs)
        except ImportError:
            with open(path, encoding="utf-8", errors="ignore") as f:
                return f.read()
    if ext in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            rows = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    rows.append(" | ".join(str(c) for c in row if c is not None))
            return "\n".join(rows)
        except Exception:
            return ""
    # Default: try plain text
    with open(path, encoding="utf-8", errors="ignore") as f:
        return f.read()


def _parse_json_array(raw: str, context: str) -> list:
    """Extract the first JSON array from an LLM response."""
    raw = raw.strip()
    # Strip markdown fences
    raw = re.sub(r"^```[a-z]*\n?", "", raw)
    raw = re.sub(r"```$", "", raw.strip())
    match = re.search(r"\[.*\]", raw, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError as e:
            print(f"[{context}] JSON parse error: {e}")
    return []

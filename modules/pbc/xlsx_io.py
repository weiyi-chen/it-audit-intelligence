"""
Excel I/O for Module A — PBC Checklist Generator.

read_pbc_xlsx(path)  → List[PBCItem]
    Reads a workbook whose first sheet has a header row followed by data rows.
    Column order is determined by header names (case-insensitive) so the
    function is robust to column reordering.  Missing optional columns get
    sensible defaults.

write_pbc_xlsx(items, path)
    Writes current_year_items back to a new workbook, applying a background
    colour to each row based on the item's status field:

        carried_over → white  (no change)
        updated      → yellow (#FFF2CC)
        new          → green  (#D9EAD3)
        removed      → red    (#F4CCCC)

    The workbook uses auto-sized columns, a bold frozen header row, and a
    thin border around every data cell for readability.
"""

from __future__ import annotations

import os
from typing import List, Optional

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from core.state import PBCItem

# ─────────────────────────────────────────────────────────────────────────────
# Colour palette (Excel ARGB hex, alpha must be "FF")
# ─────────────────────────────────────────────────────────────────────────────

_FILL: dict[str, PatternFill] = {
    "carried_over": PatternFill(fill_type=None),                          # white / no fill
    "updated":      PatternFill("solid", fgColor="FFFFF2CC"),             # soft yellow
    "new":          PatternFill("solid", fgColor="FFD9EAD3"),             # soft green
    "removed":      PatternFill("solid", fgColor="FFF4CCCC"),             # soft red
}
_DEFAULT_FILL = PatternFill(fill_type=None)

# Header row styling
_HEADER_FONT   = Font(bold=True, color="FF000000")
_HEADER_FILL   = PatternFill("solid", fgColor="FF4A86C8")  # audit-blue header
_HEADER_FONT   = Font(bold=True, color="FFFFFFFF")         # white text on blue
_THIN_SIDE     = Side(style="thin")
_THIN_BORDER   = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)

# Canonical column order for writing
_WRITE_COLUMNS = [
    "item_id",
    "category",
    "description",
    "in_scope",
    "period",
    "sample_size",
    "status",
    "last_year_id",
    "notes",
]

# Human-readable header labels
_HEADER_LABELS = {
    "item_id":      "Item ID",
    "category":     "Category",
    "description":  "Description / Evidence Request",
    "in_scope":     "In Scope",
    "period":       "Period",
    "sample_size":  "Sample Size",
    "status":       "Status",
    "last_year_id": "Prior Year ID",
    "notes":        "Notes",
}

# ─────────────────────────────────────────────────────────────────────────────
# Column-name normalisation (defined early — used at module-init time below)
# ─────────────────────────────────────────────────────────────────────────────

def _normalise_col_name(name: str) -> str:
    """'Item ID' → 'item_id',  'In Scope?' → 'in_scope'."""
    return name.strip().lower().replace(" ", "_").replace("?", "").replace("/", "_")


# Reverse map: normalised display-header string → canonical field key.
# Built at import time so read_pbc_xlsx can recognise columns written by
# write_pbc_xlsx even though their display names don't normalise to the
# field key (e.g. "Description / Evidence Request" → "description",
# "Prior Year ID" → "last_year_id").
_DISPLAY_TO_KEY: dict[str, str] = {}
for _key, _label in _HEADER_LABELS.items():
    _DISPLAY_TO_KEY[_normalise_col_name(_label)] = _key  # display header → key
    _DISPLAY_TO_KEY[_normalise_col_name(_key)]   = _key  # field key itself → key


# Approximate column widths (characters)
_COL_WIDTHS = {
    "item_id":      14,
    "category":     22,
    "description":  60,
    "in_scope":     10,
    "period":       12,
    "sample_size":  14,
    "status":       16,
    "last_year_id": 14,
    "notes":        40,
}


# ─────────────────────────────────────────────────────────────────────────────
# Read
# ─────────────────────────────────────────────────────────────────────────────

def read_pbc_xlsx(path: str) -> List[PBCItem]:
    """
    Parse a PBC xlsx workbook into a list of PBCItem dicts.

    The workbook must have at least one sheet.  Row 1 is treated as the
    header; subsequent non-empty rows become PBCItems.

    Recognised column names (case-insensitive, spaces/underscores ignored):
        item_id, category, description, in_scope, period,
        sample_size, status, last_year_id, notes

    Any unrecognised columns are silently ignored.
    Missing optional columns get defaults:
        in_scope    → True
        period      → ""
        sample_size → None
        status      → "carried_over"
        last_year_id→ None
        notes       → ""

    Raises
    ------
    FileNotFoundError  if the path doesn't exist.
    ValueError         if the workbook has no sheet or no header row.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"PBC xlsx not found: {path!r}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"Workbook has no active sheet: {path!r}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Workbook is empty: {path!r}")

    # Build field-key → column-index map from row 0.
    # Each header cell is normalised then resolved through _DISPLAY_TO_KEY so
    # that display labels like "Description / Evidence Request" or "Prior Year ID"
    # map back to the canonical field keys "description" and "last_year_id".
    col_map: dict[str, int] = {}
    for col_idx, cell_val in enumerate(rows[0]):
        if cell_val is None:
            continue
        normalised = _normalise_col_name(str(cell_val))
        field_key  = _DISPLAY_TO_KEY.get(normalised, normalised)
        col_map[field_key] = col_idx

    items: List[PBCItem] = []
    for row in rows[1:]:
        # Skip completely empty rows
        if all(v is None for v in row):
            continue

        item = PBCItem(
            item_id     = _get_str(row, col_map, "item_id",      default=""),
            category    = _get_str(row, col_map, "category",     default=""),
            description = _get_text(row, col_map, "description", default=""),
            in_scope    = _get_bool(row, col_map, "in_scope",    default=True),
            period      = _get_str(row, col_map, "period",       default=""),
            sample_size = _get_optional_str(row, col_map, "sample_size"),
            status      = _get_str(row, col_map, "status",       default="carried_over"),
            last_year_id= _get_optional_str(row, col_map, "last_year_id"),
            notes       = _get_str(row, col_map, "notes",        default=""),
        )

        # Skip rows that have no item_id and no description (header duplicates, etc.)
        if not item["item_id"] and not item["description"]:
            continue

        items.append(item)

    return items


# ─────────────────────────────────────────────────────────────────────────────
# Write
# ─────────────────────────────────────────────────────────────────────────────

def write_pbc_xlsx(items: List[PBCItem], path: str) -> None:
    """
    Serialise a list of PBCItems to an Excel workbook at *path*.

    Each row is colour-coded by status:
        carried_over → no fill (white)
        updated      → yellow
        new          → green
        removed      → red/pink

    The output directory is created automatically if it doesn't exist.
    An existing file at *path* is overwritten.
    """
    os.makedirs(os.path.dirname(path) if os.path.dirname(path) else ".", exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PBC List"

    # ── header row ─────────────────────────────────────────────────────────
    for col_idx, col_key in enumerate(_WRITE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_HEADER_LABELS.get(col_key, col_key))
        cell.font   = _HEADER_FONT
        cell.fill   = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    ws.freeze_panes = "A2"

    # ── data rows ──────────────────────────────────────────────────────────
    for row_idx, item in enumerate(items, start=2):
        fill = _FILL.get(item.get("status", "carried_over"), _DEFAULT_FILL)

        row_values = [
            item.get("item_id",      ""),
            item.get("category",     ""),
            item.get("description",  ""),
            "Yes" if item.get("in_scope", True) else "No",
            item.get("period",       ""),
            item.get("sample_size",  ""),
            item.get("status",       "carried_over"),
            item.get("last_year_id", ""),
            item.get("notes",        ""),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = fill
            cell.border = _THIN_BORDER
            col_key = _WRITE_COLUMNS[col_idx - 1]
            wrap = col_key in ("description", "notes")
            cell.alignment = Alignment(
                vertical="top", wrap_text=wrap, shrink_to_fit=not wrap
            )

    # ── column widths ──────────────────────────────────────────────────────
    for col_idx, col_key in enumerate(_WRITE_COLUMNS, start=1):
        width = _COL_WIDTHS.get(col_key, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── row height for data rows ───────────────────────────────────────────
    # Description column can wrap — give rows a bit of breathing room.
    for row_idx in range(2, len(items) + 2):
        ws.row_dimensions[row_idx].height = 30

    # ── legend sheet ───────────────────────────────────────────────────────
    legend_ws = wb.create_sheet("Legend")
    legend_data = [
        ("Status",       "Fill Colour",  "Meaning"),
        ("carried_over", "White",        "Item unchanged from prior year"),
        ("updated",      "Yellow",       "Item wording or scope updated"),
        ("new",          "Green",        "New item — added for current year"),
        ("removed",      "Red/Pink",     "Item removed (out of scope)"),
    ]
    legend_fills = {
        "carried_over": _DEFAULT_FILL,
        "updated":      _FILL["updated"],
        "new":          _FILL["new"],
        "removed":      _FILL["removed"],
    }
    for r_idx, (status_val, colour_label, meaning) in enumerate(legend_data, start=1):
        for c_idx, val in enumerate([status_val, colour_label, meaning], start=1):
            cell = legend_ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True)
            elif c_idx == 1 and status_val in legend_fills:
                cell.fill = legend_fills[status_val]
            cell.border = _THIN_BORDER
    for c_idx, width in enumerate([18, 14, 40], start=1):
        legend_ws.column_dimensions[get_column_letter(c_idx)].width = width

    wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _get_str(row: tuple, col_map: dict, key: str, default: str = "") -> str:
    normalised = _normalise_col_name(key)
    idx = col_map.get(normalised)
    if idx is None or idx >= len(row) or row[idx] is None:
        return default
    return str(row[idx]).strip()


def _get_text(row: tuple, col_map: dict, key: str, default: str = "") -> str:
    """Read free-form text without changing meaningful whitespace."""
    normalised = _normalise_col_name(key)
    idx = col_map.get(normalised)
    if idx is None or idx >= len(row) or row[idx] is None:
        return default
    return str(row[idx])


def _get_optional_str(row: tuple, col_map: dict, key: str) -> Optional[str]:
    normalised = _normalise_col_name(key)
    idx = col_map.get(normalised)
    if idx is None or idx >= len(row) or row[idx] is None:
        return None
    val = str(row[idx]).strip()
    return val if val else None


def _get_bool(row: tuple, col_map: dict, key: str, default: bool = True) -> bool:
    normalised = _normalise_col_name(key)
    idx = col_map.get(normalised)
    if idx is None or idx >= len(row) or row[idx] is None:
        return default
    val = str(row[idx]).strip().lower()
    if val in ("yes", "true", "1", "y"):
        return True
    if val in ("no", "false", "0", "n"):
        return False
    return default
